from fastapi import FastAPI, APIRouter, HTTPException, Depends, UploadFile, File
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.responses import StreamingResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
from apscheduler.schedulers.asyncio import AsyncIOScheduler
from apscheduler.triggers.cron import CronTrigger
import os
import logging
import asyncio
from pathlib import Path
from pydantic import BaseModel, Field, EmailStr
from typing import List, Optional
import uuid
from datetime import datetime, timezone, timedelta, date
import jwt
import bcrypt
import resend
import base64
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import calendar

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# JWT settings
JWT_SECRET = os.environ.get('JWT_SECRET', 'absence-management-secret-key-2024')
JWT_ALGORITHM = "HS256"
JWT_EXPIRATION_HOURS = 24

# Resend setup
resend.api_key = os.environ.get('RESEND_API_KEY', '')
SENDER_EMAIL = os.environ.get('SENDER_EMAIL', 'onboarding@resend.dev')

# Security
security = HTTPBearer()

# Create the main app
app = FastAPI(title="Absence Management API")

# Create router with /api prefix
api_router = APIRouter(prefix="/api")

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# Scheduler for monthly accrual
scheduler = AsyncIOScheduler()

# ==================== MODELS ====================

class UserRole:
    ADMIN = "admin"
    MANAGER = "manager"
    EMPLOYEE = "employee"
    HR = "ufficio_personale"

class AbsenceType:
    FERIE = "ferie"
    PERMESSO = "permesso"
    MALATTIA = "malattia"

class AbsenceStatus:
    PENDING = "pending"
    APPROVED = "approved"
    REJECTED = "rejected"

# User Models
class UserCreate(BaseModel):
    email: EmailStr
    password: str
    first_name: str
    last_name: str
    role: str = UserRole.EMPLOYEE
    manager_id: Optional[str] = None
    total_ferie_hours: float = 0
    total_permessi_hours: float = 0
    monthly_ferie_hours: float = 0
    monthly_permessi_hours: float = 0

class UserLogin(BaseModel):
    email: EmailStr
    password: str

class UserResponse(BaseModel):
    user_id: str
    email: str
    first_name: str
    last_name: str
    role: str
    manager_id: Optional[str] = None
    total_ferie_hours: float = 0
    total_permessi_hours: float = 0
    monthly_ferie_hours: float = 0
    monthly_permessi_hours: float = 0
    created_at: str

class UserUpdate(BaseModel):
    first_name: Optional[str] = None
    last_name: Optional[str] = None
    role: Optional[str] = None
    manager_id: Optional[str] = None
    total_ferie_hours: Optional[float] = None
    total_permessi_hours: Optional[float] = None
    monthly_ferie_hours: Optional[float] = None
    monthly_permessi_hours: Optional[float] = None

class AddHoursRequest(BaseModel):
    user_id: str
    hours_type: str
    hours: float
    notes: Optional[str] = None

# Absence Models
class AbsenceCreate(BaseModel):
    absence_type: str
    start_date: str
    end_date: str
    hours: Optional[float] = None
    notes: Optional[str] = None

class AbsenceResponse(BaseModel):
    absence_id: str
    user_id: str
    user_name: str
    user_email: str
    absence_type: str
    start_date: str
    end_date: str
    hours: Optional[float] = None
    status: str
    notes: Optional[str] = None
    approved_by: Optional[str] = None
    approved_at: Optional[str] = None
    rejection_reason: Optional[str] = None
    created_at: str

class AbsenceAction(BaseModel):
    action: str  # approve, reject, pending
    reason: Optional[str] = None

class AbsenceStatusChange(BaseModel):
    new_status: str  # pending, approved, rejected
    reason: Optional[str] = None

# Hours Balance Response
class HoursBalance(BaseModel):
    user_id: str
    user_name: str
    ferie_total: float
    ferie_used: float
    ferie_pending: float
    ferie_remaining: float
    permessi_total: float
    permessi_used: float
    permessi_pending: float
    permessi_remaining: float
    monthly_ferie_hours: float
    monthly_permessi_hours: float

# Settings Models
class CompanySettings(BaseModel):
    company_name: str = "My Company"
    logo_base64: Optional[str] = None

# ==================== HELPERS ====================

def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')

def verify_password(password: str, hashed: str) -> bool:
    return bcrypt.checkpw(password.encode('utf-8'), hashed.encode('utf-8'))

def create_token(user_id: str, email: str, role: str) -> str:
    payload = {
        "user_id": user_id,
        "email": email,
        "role": role,
        "exp": datetime.now(timezone.utc) + timedelta(hours=JWT_EXPIRATION_HOURS)
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)

def decode_token(token: str) -> dict:
    try:
        return jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token scaduto")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Token non valido")

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    payload = decode_token(credentials.credentials)
    user = await db.users.find_one({"user_id": payload["user_id"]}, {"_id": 0})
    if not user:
        raise HTTPException(status_code=401, detail="Utente non trovato")
    return user

async def require_admin(current_user: dict = Depends(get_current_user)):
    if current_user["role"] != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Accesso riservato agli amministratori")
    return current_user

async def require_manager_or_admin(current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in [UserRole.ADMIN, UserRole.MANAGER, UserRole.HR]:
        raise HTTPException(status_code=403, detail="Accesso riservato a manager, ufficio personale e amministratori")
    return current_user

async def require_hr_or_admin(current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in [UserRole.ADMIN, UserRole.HR]:
        raise HTTPException(status_code=403, detail="Accesso riservato a ufficio personale e amministratori")
    return current_user

async def require_hours_manager(current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in [UserRole.ADMIN, UserRole.MANAGER, UserRole.HR]:
        raise HTTPException(status_code=403, detail="Non hai i permessi per gestire le ore")
    return current_user

def calculate_working_days(start_date: str, end_date: str) -> int:
    start = datetime.strptime(start_date, "%Y-%m-%d").date()
    end = datetime.strptime(end_date, "%Y-%m-%d").date()
    days = 0
    current = start
    while current <= end:
        if current.weekday() < 5:
            days += 1
        current += timedelta(days=1)
    return days

async def calculate_user_hours_balance(user_id: str) -> dict:
    user = await db.users.find_one({"user_id": user_id}, {"_id": 0})
    if not user:
        return None
    
    approved_absences = await db.absences.find({
        "user_id": user_id,
        "status": AbsenceStatus.APPROVED
    }, {"_id": 0}).to_list(1000)
    
    pending_absences = await db.absences.find({
        "user_id": user_id,
        "status": AbsenceStatus.PENDING
    }, {"_id": 0}).to_list(1000)
    
    ferie_used = 0
    ferie_pending = 0
    permessi_used = 0
    permessi_pending = 0
    
    for absence in approved_absences:
        if absence["absence_type"] == AbsenceType.FERIE:
            days = calculate_working_days(absence["start_date"], absence["end_date"])
            ferie_used += days * 8
        elif absence["absence_type"] == AbsenceType.PERMESSO:
            permessi_used += absence.get("hours", 0)
    
    for absence in pending_absences:
        if absence["absence_type"] == AbsenceType.FERIE:
            days = calculate_working_days(absence["start_date"], absence["end_date"])
            ferie_pending += days * 8
        elif absence["absence_type"] == AbsenceType.PERMESSO:
            permessi_pending += absence.get("hours", 0)
    
    total_ferie = user.get("total_ferie_hours", 0)
    total_permessi = user.get("total_permessi_hours", 0)
    
    return {
        "user_id": user_id,
        "user_name": f"{user['first_name']} {user['last_name']}",
        "ferie_total": total_ferie,
        "ferie_used": ferie_used,
        "ferie_pending": ferie_pending,
        "ferie_remaining": total_ferie - ferie_used,
        "permessi_total": total_permessi,
        "permessi_used": permessi_used,
        "permessi_pending": permessi_pending,
        "permessi_remaining": total_permessi - permessi_used,
        "monthly_ferie_hours": user.get("monthly_ferie_hours", 0),
        "monthly_permessi_hours": user.get("monthly_permessi_hours", 0)
    }

async def send_notification_email(to_email: str, subject: str, html_content: str):
    if not resend.api_key:
        logger.warning("RESEND_API_KEY not configured, skipping email")
        return None
    
    try:
        params = {
            "from": SENDER_EMAIL,
            "to": [to_email],
            "subject": subject,
            "html": html_content
        }
        email = await asyncio.to_thread(resend.Emails.send, params)
        logger.info(f"Email sent to {to_email}")
        return email
    except Exception as e:
        logger.error(f"Failed to send email: {str(e)}")
        return None

# ==================== SCHEDULED JOBS ====================

async def monthly_accrual_job():
    """Scheduled job to run monthly accrual on the 1st of each month"""
    logger.info("Running scheduled monthly accrual job...")
    users = await db.users.find({}, {"_id": 0}).to_list(1000)
    updated_count = 0
    
    for user in users:
        monthly_ferie = user.get("monthly_ferie_hours", 0)
        monthly_permessi = user.get("monthly_permessi_hours", 0)
        
        if monthly_ferie > 0 or monthly_permessi > 0:
            update_data = {}
            if monthly_ferie > 0:
                update_data["total_ferie_hours"] = user.get("total_ferie_hours", 0) + monthly_ferie
            if monthly_permessi > 0:
                update_data["total_permessi_hours"] = user.get("total_permessi_hours", 0) + monthly_permessi
            
            if update_data:
                await db.users.update_one({"user_id": user["user_id"]}, {"$set": update_data})
                updated_count += 1
    
    await db.accrual_logs.insert_one({
        "log_id": str(uuid.uuid4()),
        "run_by": "SYSTEM",
        "run_by_name": "Maturazione Automatica",
        "users_updated": updated_count,
        "created_at": datetime.now(timezone.utc).isoformat()
    })
    
    logger.info(f"Monthly accrual completed for {updated_count} users")

# ==================== AUTH ROUTES ====================

@api_router.post("/auth/register", response_model=dict)
async def register(user: UserCreate):
    existing = await db.users.find_one({"email": user.email})
    if existing:
        raise HTTPException(status_code=400, detail="Email già registrata")
    
    user_id = str(uuid.uuid4())
    user_doc = {
        "user_id": user_id,
        "email": user.email,
        "password_hash": hash_password(user.password),
        "first_name": user.first_name,
        "last_name": user.last_name,
        "role": user.role,
        "manager_id": user.manager_id,
        "total_ferie_hours": user.total_ferie_hours,
        "total_permessi_hours": user.total_permessi_hours,
        "monthly_ferie_hours": user.monthly_ferie_hours,
        "monthly_permessi_hours": user.monthly_permessi_hours,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.users.insert_one(user_doc)
    
    token = create_token(user_id, user.email, user.role)
    return {
        "token": token,
        "user": UserResponse(
            user_id=user_id,
            email=user.email,
            first_name=user.first_name,
            last_name=user.last_name,
            role=user.role,
            manager_id=user.manager_id,
            total_ferie_hours=user.total_ferie_hours,
            total_permessi_hours=user.total_permessi_hours,
            monthly_ferie_hours=user.monthly_ferie_hours,
            monthly_permessi_hours=user.monthly_permessi_hours,
            created_at=user_doc["created_at"]
        ).model_dump()
    }

@api_router.post("/auth/login", response_model=dict)
async def login(credentials: UserLogin):
    user = await db.users.find_one({"email": credentials.email}, {"_id": 0})
    if not user or not verify_password(credentials.password, user["password_hash"]):
        raise HTTPException(status_code=401, detail="Credenziali non valide")
    
    token = create_token(user["user_id"], user["email"], user["role"])
    return {
        "token": token,
        "user": UserResponse(
            user_id=user["user_id"],
            email=user["email"],
            first_name=user["first_name"],
            last_name=user["last_name"],
            role=user["role"],
            manager_id=user.get("manager_id"),
            total_ferie_hours=user.get("total_ferie_hours", 0),
            total_permessi_hours=user.get("total_permessi_hours", 0),
            monthly_ferie_hours=user.get("monthly_ferie_hours", 0),
            monthly_permessi_hours=user.get("monthly_permessi_hours", 0),
            created_at=user["created_at"]
        ).model_dump()
    }

@api_router.get("/auth/me", response_model=UserResponse)
async def get_me(current_user: dict = Depends(get_current_user)):
    return UserResponse(
        user_id=current_user["user_id"],
        email=current_user["email"],
        first_name=current_user["first_name"],
        last_name=current_user["last_name"],
        role=current_user["role"],
        manager_id=current_user.get("manager_id"),
        total_ferie_hours=current_user.get("total_ferie_hours", 0),
        total_permessi_hours=current_user.get("total_permessi_hours", 0),
        monthly_ferie_hours=current_user.get("monthly_ferie_hours", 0),
        monthly_permessi_hours=current_user.get("monthly_permessi_hours", 0),
        created_at=current_user["created_at"]
    )

# ==================== USER ROUTES ====================

@api_router.get("/users", response_model=List[UserResponse])
async def get_users(current_user: dict = Depends(require_manager_or_admin)):
    users = await db.users.find({}, {"_id": 0, "password_hash": 0}).to_list(1000)
    return [UserResponse(
        user_id=u["user_id"],
        email=u["email"],
        first_name=u["first_name"],
        last_name=u["last_name"],
        role=u["role"],
        manager_id=u.get("manager_id"),
        total_ferie_hours=u.get("total_ferie_hours", 0),
        total_permessi_hours=u.get("total_permessi_hours", 0),
        monthly_ferie_hours=u.get("monthly_ferie_hours", 0),
        monthly_permessi_hours=u.get("monthly_permessi_hours", 0),
        created_at=u["created_at"]
    ) for u in users]

@api_router.get("/users/{user_id}", response_model=UserResponse)
async def get_user(user_id: str, current_user: dict = Depends(get_current_user)):
    if current_user["user_id"] != user_id and current_user["role"] not in [UserRole.ADMIN, UserRole.MANAGER, UserRole.HR]:
        raise HTTPException(status_code=403, detail="Non autorizzato")
    
    user = await db.users.find_one({"user_id": user_id}, {"_id": 0, "password_hash": 0})
    if not user:
        raise HTTPException(status_code=404, detail="Utente non trovato")
    return UserResponse(
        user_id=user["user_id"],
        email=user["email"],
        first_name=user["first_name"],
        last_name=user["last_name"],
        role=user["role"],
        manager_id=user.get("manager_id"),
        total_ferie_hours=user.get("total_ferie_hours", 0),
        total_permessi_hours=user.get("total_permessi_hours", 0),
        monthly_ferie_hours=user.get("monthly_ferie_hours", 0),
        monthly_permessi_hours=user.get("monthly_permessi_hours", 0),
        created_at=user["created_at"]
    )

@api_router.put("/users/{user_id}", response_model=UserResponse)
async def update_user(user_id: str, update: UserUpdate, current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in [UserRole.ADMIN, UserRole.HR]:
        raise HTTPException(status_code=403, detail="Non autorizzato")
    
    user = await db.users.find_one({"user_id": user_id}, {"_id": 0})
    if not user:
        raise HTTPException(status_code=404, detail="Utente non trovato")
    
    update_data = {k: v for k, v in update.model_dump().items() if v is not None}
    
    if current_user["role"] == UserRole.HR:
        allowed_fields = ["total_ferie_hours", "total_permessi_hours", "monthly_ferie_hours", "monthly_permessi_hours"]
        update_data = {k: v for k, v in update_data.items() if k in allowed_fields}
    
    if update_data:
        await db.users.update_one({"user_id": user_id}, {"$set": update_data})
    
    updated_user = await db.users.find_one({"user_id": user_id}, {"_id": 0, "password_hash": 0})
    return UserResponse(
        user_id=updated_user["user_id"],
        email=updated_user["email"],
        first_name=updated_user["first_name"],
        last_name=updated_user["last_name"],
        role=updated_user["role"],
        manager_id=updated_user.get("manager_id"),
        total_ferie_hours=updated_user.get("total_ferie_hours", 0),
        total_permessi_hours=updated_user.get("total_permessi_hours", 0),
        monthly_ferie_hours=updated_user.get("monthly_ferie_hours", 0),
        monthly_permessi_hours=updated_user.get("monthly_permessi_hours", 0),
        created_at=updated_user["created_at"]
    )

@api_router.delete("/users/{user_id}")
async def delete_user(user_id: str, current_user: dict = Depends(require_admin)):
    if user_id == current_user["user_id"]:
        raise HTTPException(status_code=400, detail="Non puoi eliminare te stesso")
    
    result = await db.users.delete_one({"user_id": user_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Utente non trovato")
    
    await db.users.update_many({"manager_id": user_id}, {"$set": {"manager_id": None}})
    return {"message": "Utente eliminato"}

# ==================== HOURS MANAGEMENT ROUTES ====================

@api_router.get("/users/{user_id}/balance", response_model=HoursBalance)
async def get_user_balance(user_id: str, current_user: dict = Depends(get_current_user)):
    if current_user["user_id"] != user_id and current_user["role"] not in [UserRole.ADMIN, UserRole.MANAGER, UserRole.HR]:
        raise HTTPException(status_code=403, detail="Non autorizzato")
    
    balance = await calculate_user_hours_balance(user_id)
    if not balance:
        raise HTTPException(status_code=404, detail="Utente non trovato")
    return HoursBalance(**balance)

@api_router.get("/balance/my", response_model=HoursBalance)
async def get_my_balance(current_user: dict = Depends(get_current_user)):
    balance = await calculate_user_hours_balance(current_user["user_id"])
    return HoursBalance(**balance)

@api_router.get("/balance/all", response_model=List[HoursBalance])
async def get_all_balances(current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in [UserRole.ADMIN, UserRole.MANAGER, UserRole.HR]:
        raise HTTPException(status_code=403, detail="Non autorizzato")
    
    users = await db.users.find({}, {"_id": 0, "user_id": 1}).to_list(1000)
    balances = []
    for user in users:
        balance = await calculate_user_hours_balance(user["user_id"])
        if balance:
            balances.append(HoursBalance(**balance))
    return balances

@api_router.post("/users/{user_id}/add-hours")
async def add_hours_to_user(user_id: str, request: AddHoursRequest, current_user: dict = Depends(require_hours_manager)):
    user = await db.users.find_one({"user_id": user_id}, {"_id": 0})
    if not user:
        raise HTTPException(status_code=404, detail="Utente non trovato")
    
    if request.hours_type == "ferie":
        new_total = user.get("total_ferie_hours", 0) + request.hours
        await db.users.update_one({"user_id": user_id}, {"$set": {"total_ferie_hours": new_total}})
    elif request.hours_type == "permessi":
        new_total = user.get("total_permessi_hours", 0) + request.hours
        await db.users.update_one({"user_id": user_id}, {"$set": {"total_permessi_hours": new_total}})
    else:
        raise HTTPException(status_code=400, detail="Tipo ore non valido")
    
    await db.hours_adjustments.insert_one({
        "adjustment_id": str(uuid.uuid4()),
        "user_id": user_id,
        "hours_type": request.hours_type,
        "hours": request.hours,
        "notes": request.notes,
        "adjusted_by": current_user["user_id"],
        "adjusted_by_name": f"{current_user['first_name']} {current_user['last_name']}",
        "created_at": datetime.now(timezone.utc).isoformat()
    })
    
    return {"message": f"Aggiunte {request.hours} ore di {request.hours_type}", "new_total": new_total}

@api_router.post("/hours/monthly-accrual")
async def run_monthly_accrual(current_user: dict = Depends(require_hr_or_admin)):
    users = await db.users.find({}, {"_id": 0}).to_list(1000)
    updated_count = 0
    
    for user in users:
        monthly_ferie = user.get("monthly_ferie_hours", 0)
        monthly_permessi = user.get("monthly_permessi_hours", 0)
        
        if monthly_ferie > 0 or monthly_permessi > 0:
            update_data = {}
            if monthly_ferie > 0:
                update_data["total_ferie_hours"] = user.get("total_ferie_hours", 0) + monthly_ferie
            if monthly_permessi > 0:
                update_data["total_permessi_hours"] = user.get("total_permessi_hours", 0) + monthly_permessi
            
            if update_data:
                await db.users.update_one({"user_id": user["user_id"]}, {"$set": update_data})
                updated_count += 1
    
    await db.accrual_logs.insert_one({
        "log_id": str(uuid.uuid4()),
        "run_by": current_user["user_id"],
        "run_by_name": f"{current_user['first_name']} {current_user['last_name']}",
        "users_updated": updated_count,
        "created_at": datetime.now(timezone.utc).isoformat()
    })
    
    return {"message": f"Maturazione mensile completata per {updated_count} utenti"}

# ==================== ABSENCE ROUTES ====================

@api_router.post("/absences", response_model=AbsenceResponse)
async def create_absence(absence: AbsenceCreate, current_user: dict = Depends(get_current_user)):
    if absence.absence_type == AbsenceType.PERMESSO and not absence.hours:
        raise HTTPException(status_code=400, detail="Le ore sono obbligatorie per i permessi")
    
    absence_id = str(uuid.uuid4())
    absence_doc = {
        "absence_id": absence_id,
        "user_id": current_user["user_id"],
        "user_name": f"{current_user['first_name']} {current_user['last_name']}",
        "user_email": current_user["email"],
        "absence_type": absence.absence_type,
        "start_date": absence.start_date,
        "end_date": absence.end_date,
        "hours": absence.hours if absence.absence_type == AbsenceType.PERMESSO else None,
        "status": AbsenceStatus.PENDING,
        "notes": absence.notes,
        "approved_by": None,
        "approved_at": None,
        "rejection_reason": None,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.absences.insert_one(absence_doc)
    
    managers = await db.users.find(
        {"$or": [
            {"role": UserRole.ADMIN},
            {"role": UserRole.HR},
            {"role": UserRole.MANAGER, "user_id": current_user.get("manager_id")}
        ]}, 
        {"_id": 0, "email": 1}
    ).to_list(100)
    
    absence_type_labels = {"ferie": "Ferie", "permesso": "Permesso", "malattia": "Malattia"}
    hours_info = f"<p><strong>Ore:</strong> {absence.hours}</p>" if absence.hours else ""
    
    for manager in managers:
        await send_notification_email(
            manager["email"],
            f"Nuova richiesta di {absence_type_labels.get(absence.absence_type, absence.absence_type)}",
            f"""
            <h2>Nuova richiesta di assenza</h2>
            <p><strong>Dipendente:</strong> {current_user['first_name']} {current_user['last_name']}</p>
            <p><strong>Tipo:</strong> {absence_type_labels.get(absence.absence_type, absence.absence_type)}</p>
            <p><strong>Dal:</strong> {absence.start_date}</p>
            <p><strong>Al:</strong> {absence.end_date}</p>
            {hours_info}
            <p><strong>Note:</strong> {absence.notes or 'Nessuna'}</p>
            """
        )
    
    return AbsenceResponse(**{k: v for k, v in absence_doc.items() if k != "_id"})

@api_router.get("/absences", response_model=List[AbsenceResponse])
async def get_absences(
    absence_type: Optional[str] = None,
    status: Optional[str] = None,
    user_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    query = {}
    if absence_type:
        query["absence_type"] = absence_type
    if status:
        query["status"] = status
    if user_id:
        query["user_id"] = user_id
    
    absences = await db.absences.find(query, {"_id": 0}).to_list(1000)
    return [AbsenceResponse(**a) for a in absences]

@api_router.get("/absences/my", response_model=List[AbsenceResponse])
async def get_my_absences(current_user: dict = Depends(get_current_user)):
    absences = await db.absences.find({"user_id": current_user["user_id"]}, {"_id": 0}).to_list(100)
    return [AbsenceResponse(**a) for a in absences]

@api_router.get("/absences/pending", response_model=List[AbsenceResponse])
async def get_pending_absences(current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in [UserRole.ADMIN, UserRole.MANAGER, UserRole.HR]:
        raise HTTPException(status_code=403, detail="Non autorizzato")
    
    query = {"status": AbsenceStatus.PENDING}
    
    if current_user["role"] == UserRole.MANAGER:
        team_members = await db.users.find(
            {"manager_id": current_user["user_id"]}, 
            {"_id": 0, "user_id": 1}
        ).to_list(100)
        team_ids = [m["user_id"] for m in team_members]
        query["user_id"] = {"$in": team_ids}
    
    absences = await db.absences.find(query, {"_id": 0}).to_list(100)
    return [AbsenceResponse(**a) for a in absences]

@api_router.put("/absences/{absence_id}/action", response_model=AbsenceResponse)
async def handle_absence_action(
    absence_id: str, 
    action: AbsenceAction, 
    current_user: dict = Depends(get_current_user)
):
    if current_user["role"] not in [UserRole.ADMIN, UserRole.MANAGER, UserRole.HR]:
        raise HTTPException(status_code=403, detail="Non autorizzato")
    
    absence = await db.absences.find_one({"absence_id": absence_id}, {"_id": 0})
    if not absence:
        raise HTTPException(status_code=404, detail="Richiesta non trovata")
    
    if current_user["role"] == UserRole.MANAGER:
        requester = await db.users.find_one({"user_id": absence["user_id"]}, {"_id": 0})
        if requester and requester.get("manager_id") != current_user["user_id"]:
            raise HTTPException(status_code=403, detail="Non puoi gestire richieste di altri team")
    
    absence_type_labels = {"ferie": "Ferie", "permesso": "Permesso", "malattia": "Malattia"}
    
    if action.action == "approve":
        new_status = AbsenceStatus.APPROVED
        email_subject = "Richiesta approvata"
        email_message = "La tua richiesta di assenza è stata <strong>APPROVATA</strong>."
    elif action.action == "reject":
        new_status = AbsenceStatus.REJECTED
        email_subject = "Richiesta rifiutata"
        email_message = f"La tua richiesta di assenza è stata <strong>RIFIUTATA</strong>.<br>Motivo: {action.reason or 'Non specificato'}"
    elif action.action == "pending":
        new_status = AbsenceStatus.PENDING
        email_subject = "Richiesta rimessa in attesa"
        email_message = "La tua richiesta di assenza è stata <strong>RIMESSA IN ATTESA</strong> di revisione."
    else:
        raise HTTPException(status_code=400, detail="Azione non valida. Usa: approve, reject, pending")
    
    update_data = {
        "status": new_status,
        "approved_by": f"{current_user['first_name']} {current_user['last_name']}",
        "approved_at": datetime.now(timezone.utc).isoformat(),
        "rejection_reason": action.reason if action.action == "reject" else None
    }
    
    await db.absences.update_one({"absence_id": absence_id}, {"$set": update_data})
    
    await send_notification_email(
        absence["user_email"],
        f"{email_subject} - {absence_type_labels.get(absence['absence_type'], absence['absence_type'])}",
        f"""
        <h2>{email_subject}</h2>
        <p>{email_message}</p>
        <p><strong>Tipo:</strong> {absence_type_labels.get(absence['absence_type'], absence['absence_type'])}</p>
        <p><strong>Dal:</strong> {absence['start_date']}</p>
        <p><strong>Al:</strong> {absence['end_date']}</p>
        <p><strong>Gestito da:</strong> {current_user['first_name']} {current_user['last_name']}</p>
        """
    )
    
    updated = await db.absences.find_one({"absence_id": absence_id}, {"_id": 0})
    return AbsenceResponse(**updated)

@api_router.put("/absences/{absence_id}/status", response_model=AbsenceResponse)
async def change_absence_status(
    absence_id: str, 
    status_change: AbsenceStatusChange, 
    current_user: dict = Depends(get_current_user)
):
    """Change status of an already processed absence - for admin/manager/hr"""
    if current_user["role"] not in [UserRole.ADMIN, UserRole.MANAGER, UserRole.HR]:
        raise HTTPException(status_code=403, detail="Non autorizzato")
    
    absence = await db.absences.find_one({"absence_id": absence_id}, {"_id": 0})
    if not absence:
        raise HTTPException(status_code=404, detail="Richiesta non trovata")
    
    if status_change.new_status not in [AbsenceStatus.PENDING, AbsenceStatus.APPROVED, AbsenceStatus.REJECTED]:
        raise HTTPException(status_code=400, detail="Stato non valido. Usa: pending, approved, rejected")
    
    if current_user["role"] == UserRole.MANAGER:
        requester = await db.users.find_one({"user_id": absence["user_id"]}, {"_id": 0})
        if requester and requester.get("manager_id") != current_user["user_id"]:
            raise HTTPException(status_code=403, detail="Non puoi gestire richieste di altri team")
    
    update_data = {
        "status": status_change.new_status,
        "approved_by": f"{current_user['first_name']} {current_user['last_name']}",
        "approved_at": datetime.now(timezone.utc).isoformat(),
        "rejection_reason": status_change.reason if status_change.new_status == AbsenceStatus.REJECTED else None
    }
    
    await db.absences.update_one({"absence_id": absence_id}, {"$set": update_data})
    
    status_labels = {"pending": "In attesa", "approved": "Approvata", "rejected": "Rifiutata"}
    absence_type_labels = {"ferie": "Ferie", "permesso": "Permesso", "malattia": "Malattia"}
    
    await send_notification_email(
        absence["user_email"],
        f"Stato richiesta modificato - {absence_type_labels.get(absence['absence_type'], absence['absence_type'])}",
        f"""
        <h2>Stato richiesta modificato</h2>
        <p>Lo stato della tua richiesta è stato cambiato in: <strong>{status_labels.get(status_change.new_status, status_change.new_status).upper()}</strong></p>
        <p><strong>Tipo:</strong> {absence_type_labels.get(absence['absence_type'], absence['absence_type'])}</p>
        <p><strong>Dal:</strong> {absence['start_date']}</p>
        <p><strong>Al:</strong> {absence['end_date']}</p>
        <p><strong>Modificato da:</strong> {current_user['first_name']} {current_user['last_name']}</p>
        {f"<p><strong>Motivo:</strong> {status_change.reason}</p>" if status_change.reason else ""}
        """
    )
    
    updated = await db.absences.find_one({"absence_id": absence_id}, {"_id": 0})
    return AbsenceResponse(**updated)

@api_router.delete("/absences/{absence_id}")
async def delete_absence(absence_id: str, current_user: dict = Depends(get_current_user)):
    absence = await db.absences.find_one({"absence_id": absence_id}, {"_id": 0})
    if not absence:
        raise HTTPException(status_code=404, detail="Richiesta non trovata")
    
    if current_user["role"] != UserRole.ADMIN:
        if absence["user_id"] != current_user["user_id"]:
            raise HTTPException(status_code=403, detail="Non puoi eliminare richieste di altri")
        if absence["status"] != AbsenceStatus.PENDING:
            raise HTTPException(status_code=400, detail="Puoi eliminare solo richieste in attesa")
    
    await db.absences.delete_one({"absence_id": absence_id})
    return {"message": "Richiesta eliminata"}

# ==================== EXPORT ROUTES ====================

@api_router.get("/export/absences")
async def export_absences(
    year: int = None,
    month: int = None,
    user_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Export absences to Excel. Admin/Manager/HR can export all, employees only their own."""
    can_export_all = current_user["role"] in [UserRole.ADMIN, UserRole.MANAGER, UserRole.HR]
    
    # If user_id specified, check permissions
    if user_id and user_id != current_user["user_id"] and not can_export_all:
        raise HTTPException(status_code=403, detail="Non puoi esportare dati di altri utenti")
    
    # Default to current year/month if not specified
    now = datetime.now()
    year = year or now.year
    month = month or now.month
    
    # Build query
    query = {"status": AbsenceStatus.APPROVED}
    
    if not can_export_all or (user_id and user_id == current_user["user_id"]):
        query["user_id"] = current_user["user_id"]
    elif user_id:
        query["user_id"] = user_id
    
    # Get absences
    absences = await db.absences.find(query, {"_id": 0}).to_list(1000)
    
    # Get users for the export
    if can_export_all and not user_id:
        users = await db.users.find({}, {"_id": 0, "user_id": 1, "first_name": 1, "last_name": 1}).to_list(1000)
    else:
        target_user_id = user_id or current_user["user_id"]
        users = await db.users.find({"user_id": target_user_id}, {"_id": 0, "user_id": 1, "first_name": 1, "last_name": 1}).to_list(1)
    
    # Create workbook
    wb = Workbook()
    
    # Create sheets for each month
    days_in_month = calendar.monthrange(year, month)[1]
    month_name = calendar.month_name[month]
    
    ws = wb.active
    ws.title = f"{month_name} {year}"
    
    # Styles
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    ferie_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    permesso_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    malattia_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Header row with day initials
    day_initials = ["L", "M", "M", "G", "V", "S", "D"]  # Italian day initials
    ws.cell(row=1, column=1, value="Dipendente").font = header_font
    ws.cell(row=1, column=1).fill = header_fill
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=1, column=1).border = thin_border
    ws.column_dimensions['A'].width = 25
    
    for day in range(1, days_in_month + 1):
        col = day + 1
        day_date = date(year, month, day)
        day_initial = day_initials[day_date.weekday()]
        
        # Header with day initial
        cell = ws.cell(row=1, column=col, value=day_initial)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
        
        # Second row with day number
        cell2 = ws.cell(row=2, column=col, value=day)
        cell2.font = Font(bold=True, size=9)
        cell2.alignment = Alignment(horizontal="center", vertical="center")
        cell2.border = thin_border
        
        ws.column_dimensions[get_column_letter(col)].width = 4
    
    # Employee name header spans row 1-2
    ws.merge_cells('A1:A2')
    
    # Add legend columns
    legend_col = days_in_month + 3
    ws.cell(row=1, column=legend_col, value="Legenda:").font = Font(bold=True)
    ws.cell(row=2, column=legend_col, value="F = Ferie").fill = ferie_fill
    ws.cell(row=3, column=legend_col, value="P = Permesso").fill = permesso_fill
    ws.cell(row=4, column=legend_col, value="M = Malattia").fill = malattia_fill
    
    # Data rows
    row_num = 3
    for user in users:
        user_name = f"{user['first_name']} {user['last_name']}"
        ws.cell(row=row_num, column=1, value=user_name).border = thin_border
        ws.cell(row=row_num, column=1).alignment = Alignment(vertical="center")
        
        # Get user's absences for this month
        user_absences = [a for a in absences if a["user_id"] == user["user_id"]]
        
        for day in range(1, days_in_month + 1):
            col = day + 1
            cell = ws.cell(row=row_num, column=col)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
            current_date = date(year, month, day)
            
            # Check if this day has any absence
            for absence in user_absences:
                start = datetime.strptime(absence["start_date"], "%Y-%m-%d").date()
                end = datetime.strptime(absence["end_date"], "%Y-%m-%d").date()
                
                if start <= current_date <= end:
                    if absence["absence_type"] == "ferie":
                        cell.value = "F"
                        cell.fill = ferie_fill
                    elif absence["absence_type"] == "permesso":
                        cell.value = "P"
                        cell.fill = permesso_fill
                    elif absence["absence_type"] == "malattia":
                        cell.value = "M"
                        cell.fill = malattia_fill
                    break
        
        row_num += 1
    
    # Summary sheet
    ws_summary = wb.create_sheet(title="Riepilogo")
    ws_summary.cell(row=1, column=1, value="Dipendente").font = header_font
    ws_summary.cell(row=1, column=1).fill = header_fill
    ws_summary.cell(row=1, column=2, value="Giorni Ferie").font = header_font
    ws_summary.cell(row=1, column=2).fill = header_fill
    ws_summary.cell(row=1, column=3, value="Ore Permessi").font = header_font
    ws_summary.cell(row=1, column=3).fill = header_fill
    ws_summary.cell(row=1, column=4, value="Giorni Malattia").font = header_font
    ws_summary.cell(row=1, column=4).fill = header_fill
    
    ws_summary.column_dimensions['A'].width = 25
    ws_summary.column_dimensions['B'].width = 15
    ws_summary.column_dimensions['C'].width = 15
    ws_summary.column_dimensions['D'].width = 15
    
    row_num = 2
    for user in users:
        user_name = f"{user['first_name']} {user['last_name']}"
        user_absences = [a for a in absences if a["user_id"] == user["user_id"]]
        
        ferie_days = 0
        permesso_hours = 0
        malattia_days = 0
        
        for absence in user_absences:
            start = datetime.strptime(absence["start_date"], "%Y-%m-%d").date()
            end = datetime.strptime(absence["end_date"], "%Y-%m-%d").date()
            
            # Count only days in the selected month
            current = start
            while current <= end:
                if current.year == year and current.month == month:
                    if current.weekday() < 5:  # Working days only
                        if absence["absence_type"] == "ferie":
                            ferie_days += 1
                        elif absence["absence_type"] == "malattia":
                            malattia_days += 1
                current += timedelta(days=1)
            
            if absence["absence_type"] == "permesso":
                permesso_hours += absence.get("hours", 0)
        
        ws_summary.cell(row=row_num, column=1, value=user_name).border = thin_border
        ws_summary.cell(row=row_num, column=2, value=ferie_days).border = thin_border
        ws_summary.cell(row=row_num, column=3, value=permesso_hours).border = thin_border
        ws_summary.cell(row=row_num, column=4, value=malattia_days).border = thin_border
        row_num += 1
    
    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    filename = f"assenze_{month_name}_{year}.xlsx"
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

# ==================== SETTINGS ROUTES ====================

@api_router.get("/settings", response_model=CompanySettings)
async def get_settings(current_user: dict = Depends(get_current_user)):
    settings = await db.settings.find_one({"setting_id": "company"}, {"_id": 0})
    if not settings:
        return CompanySettings()
    return CompanySettings(
        company_name=settings.get("company_name", "My Company"),
        logo_base64=settings.get("logo_base64")
    )

@api_router.put("/settings", response_model=CompanySettings)
async def update_settings(settings: CompanySettings, current_user: dict = Depends(require_admin)):
    await db.settings.update_one(
        {"setting_id": "company"},
        {"$set": {
            "setting_id": "company",
            "company_name": settings.company_name,
            "logo_base64": settings.logo_base64
        }},
        upsert=True
    )
    return settings

@api_router.post("/settings/logo")
async def upload_logo(file: UploadFile = File(...), current_user: dict = Depends(require_admin)):
    if not file.content_type.startswith("image/"):
        raise HTTPException(status_code=400, detail="Il file deve essere un'immagine")
    
    contents = await file.read()
    if len(contents) > 2 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="Il file non può superare 2MB")
    
    logo_base64 = base64.b64encode(contents).decode('utf-8')
    content_type = file.content_type
    data_uri = f"data:{content_type};base64,{logo_base64}"
    
    await db.settings.update_one(
        {"setting_id": "company"},
        {"$set": {"logo_base64": data_uri}},
        upsert=True
    )
    
    return {"logo_base64": data_uri}

# ==================== SCHEDULER SETTINGS ====================

@api_router.get("/scheduler/status")
async def get_scheduler_status(current_user: dict = Depends(require_hr_or_admin)):
    """Get scheduler status"""
    jobs = scheduler.get_jobs()
    return {
        "running": scheduler.running,
        "jobs": [{"id": job.id, "next_run": str(job.next_run_time)} for job in jobs]
    }

# ==================== STATS ROUTES ====================

@api_router.get("/stats")
async def get_stats(current_user: dict = Depends(get_current_user)):
    total_users = await db.users.count_documents({})
    total_absences = await db.absences.count_documents({})
    pending_absences = await db.absences.count_documents({"status": AbsenceStatus.PENDING})
    approved_absences = await db.absences.count_documents({"status": AbsenceStatus.APPROVED})
    
    by_type = {}
    for atype in [AbsenceType.FERIE, AbsenceType.PERMESSO, AbsenceType.MALATTIA]:
        by_type[atype] = await db.absences.count_documents({"absence_type": atype, "status": AbsenceStatus.APPROVED})
    
    return {
        "total_users": total_users,
        "total_absences": total_absences,
        "pending_absences": pending_absences,
        "approved_absences": approved_absences,
        "by_type": by_type
    }

# ==================== HEALTH CHECK ====================

@api_router.get("/")
async def root():
    return {"message": "Absence Management API", "status": "online"}

# Include router
app.include_router(api_router)

# CORS middleware
app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.on_event("startup")
async def startup_event():
    # Schedule monthly accrual job - runs on the 1st of each month at 00:01
    scheduler.add_job(
        monthly_accrual_job,
        CronTrigger(day=1, hour=0, minute=1),
        id="monthly_accrual",
        replace_existing=True
    )
    scheduler.start()
    logger.info("Scheduler started - Monthly accrual scheduled for 1st of each month at 00:01")

@app.on_event("shutdown")
async def shutdown_event():
    scheduler.shutdown()
    client.close()
